"""정비사업 Phase 1 수집기 — 循環 Ⅰ장 「공급」의 원천(재개발·재건축 파이프라인).

실행: /usr/bin/python3 src/collect/jeongbi.py [--refresh]
  * requests·openpyxl 필요 → macOS CommandLineTools python3(/usr/bin/python3)에 존재.
    (주의: /Users/iseul/개발/venv 에는 openpyxl 없음 — 반드시 /usr/bin/python3 로 실행.)
  * 구식 .xls 를 받게 되면 soffice --headless --convert-to xlsx 폴백(_ensure_xlsx).
  * --refresh 없으면 data/raw/jeongbi/ 캐시 재사용(무네트워크 재실행 가능).

산출:
  data/raw/jeongbi/            원본 캐시(.gitignore): daegu_jeongbi_YYYY-MM.xlsx · gyeonggi_sheet.csv · busan_jeongbi.json
  data/jeongbi_phase1.json     정규화 스키마(regions 3개 지역: 대구·경기·부산)

──────────────────────────────────────────────────────────────────────────────
취득 경로 (2026-07 확정. data.go.kr fileData 는 로그인 필수화되어 원 출처로 우회)
──────────────────────────────────────────────────────────────────────────────
[대구] data.go.kr 15141940 "정비사업 추진현황"(전국 표준 9단계 템플릿)
  ✗ data.go.kr fileData: POST /tcs/dss/selectFileDataDownload.do 는 status=true 를 주지만
     atchFileId·fileDetailSn 을 비로그인 세션에는 null 로 내려줌 → 파일 취득 불가(로그인 필수 확인).
  ✓ 원 출처 daegu.go.kr/build 자료실(menu_id=00001338, bbsId=BBS_00153)에서 무로그인 취득:
     list(searchWrd=정비사업 추진현황) → 최신 '정비사업 추진현황(YYYY.M월말)' 게시글 nttId
     → POST /icms/bbs/selectBoardArticle.do → 본문의 fn_egov_downFile('FILE_...','0')
     → GET /icms/cmm/fms/FileDown.do?atchFileId=..&fileSn=0   (.xlsx, PK zip)
     ⚠ 함정: 접두 없는 /cmm/fms/FileDown.do 는 한글 파일명 인코딩 버그로 "Could not get file name"
        HTML(205B) 반환 → 반드시 /icms/ 접두 컨트롤러 사용. 세션쿠키+Referer(자료실) 필수.
     ⚠ 원본은 9단계 외 예비평가·안전진단(재건축) 2단계를 추가로 보유(년/월/일 3열 구조).

[경기] data.go.kr 15119846 "일반 정비 사업 추진 현황"(일간 CSV)
  ✓ 경기데이터드림 Sheet 직다운로드(무로그인):
     GET https://data.gg.go.kr/portal/data/sheet/downloadSheetData.do
         ?infId=S62GFEEN7JMLMA0PH6CF19108891&infSeq=1&downloadType=C   (CP949 CSV, 53열)
     서비스명/URL 은 gimi9.com 카탈로그로 확정. (대안: openapi.gg.go.kr + config.gg_key)
  ⚠ 함정: col 'FF(50) 현추진상황' 은 전부 공란 → 스냅샷 status 는 col 'B(1) 사업단계' 사용.

[부산] data.go.kr 3069406 "정비사업 정보"(활용신청 완료 → service_key 로 호출)
  ✓ GET https://apis.data.go.kr/6260000/MaintenanceBusinessStatus1/getMaintenanceBusiness1
        ?serviceKey=<service_key>&pageNo=1&numOfRows=1000&resultType=json
     스냅샷 340건(step=추진단계, 이력 인가일자 없음) → dates 전부 null, snapshot_only=true.

──────────────────────────────────────────────────────────────────────────────
함정(정찰 확정)
  · data.go.kr fileData 다운로드는 2024+ 로그인 필수(atchFileId 비공개) → 원 출처 우회.
  · 소규모주택정비(가로주택·소규모재건축·소규모재개발 = 빈집및소규모주택정비특례법, 別法)는
    도정법 정비사업과 혼합 금지 → small_scale 플래그로 분리(제거 아님).
    부산 340건 중 108건이 소규모(가로주택55·소규모재건축49·소규모재개발4).
    대구/경기 원본은 소규모를 별도 게시글/데이터셋으로 분리 제공 → 본 3개 소스는 오염 없음.
  · 스냅샷(status)과 이력(dates: 각 인가일)을 혼동 금지.
"""

import argparse
import datetime as _dt
import json
import re
import subprocess
import sys
from pathlib import Path

try:
    import requests
    import openpyxl
except ImportError as e:  # 오류 삼킴 금지 — 인터프리터 안내 후 중단
    sys.exit(f"[jeongbi] 의존성 없음({e}). /usr/bin/python3 로 실행하세요 "
             f"(requests·openpyxl 보유). 현재: {sys.executable}")

ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = ROOT / "data" / "raw" / "jeongbi"
OUT = ROOT / "data" / "jeongbi_phase1.json"
CONFIG = ROOT / "config.json"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120 Safari/537.36")

# 정규화 스키마 — 전국 표준 9단계(대구 템플릿 기준)
SCHEMA = ["예정구역고시", "구역지정", "추진위", "조합설립",
          "사업시행", "관리처분", "착공", "일반분양", "준공"]

DAEGU_LIST = ("https://www.daegu.go.kr/build/index.do"
              "?menu_id=00001338&searchCnd=0&searchWrd=%EC%A0%95%EB%B9%84%EC%82%AC%EC%97%85%20%EC%B6%94%EC%A7%84%ED%98%84%ED%99%A9")
DAEGU_BBS = {"bbsId": "BBS_00153", "bbsTyCode": "BBST01",
             "bbsAttrbCode": "BBSA03", "menu_id": "00001338"}
GG_CSV = ("https://data.gg.go.kr/portal/data/sheet/downloadSheetData.do"
          "?infId=S62GFEEN7JMLMA0PH6CF19108891&infSeq=1&downloadType=C")
GG_REFERER = ("https://data.gg.go.kr/portal/data/service/selectServicePage.do"
              "?infId=S62GFEEN7JMLMA0PH6CF19108891&infSeq=1")
BUSAN_API = ("https://apis.data.go.kr/6260000/MaintenanceBusinessStatus1/"
             "getMaintenanceBusiness1")


# ─────────────────────────── 파싱 헬퍼 ───────────────────────────
def _clean(x):
    if x is None:
        return ""
    s = str(x).strip()
    if re.fullmatch(r"-?\d+\.0", s):   # '2006.0' → '2006'
        s = s[:-2]
    return s


def _valid_ymd(y, m, d):
    try:
        y, m, d = int(y), int(m), int(d)
    except (TypeError, ValueError):
        return None
    if not (1900 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31):
        return None
    try:
        _dt.date(y, m, d)
    except ValueError:
        return None
    return f"{y:04d}-{m:02d}-{d:02d}"


def parse_ymd(y, m, d):
    """대구: 분리된 년/월/일 셀 → ('YYYY-MM-DD'|None, 원문|None).
    완전한 날짜만 성공. 부분(연도만 등)/무효는 원문 보존."""
    y, m, d = _clean(y), _clean(m), _clean(d)
    if not any([y, m, d]):
        return None, None
    iso = _valid_ymd(y, m, d)
    if iso:
        return iso, None
    return None, ".".join(p for p in (y, m, d) if p)   # 부분 날짜 원문


def parse_datestr(raw):
    """경기: 'YYYYMMDD'/'YYYY.MM.DD' 등 문자열 → (iso|None, 원문|None)."""
    s = _clean(raw)
    if not s:
        return None, None
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        iso = _valid_ymd(digits[:4], digits[4:6], digits[6:8])
        if iso:
            return iso, None
    m = re.match(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        iso = _valid_ymd(*m.groups())
        if iso:
            return iso, None
    return None, s   # 파싱 불가 → 원문 보존


def parse_float(x):
    s = _clean(x).replace(",", "")
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None


def parse_int(x, zero_null=False):
    f = parse_float(x)
    if f is None:
        return None
    i = int(round(f))
    return None if (zero_null and i == 0) else i


def classify(raw):
    """사업유형/구역명 → (type[재개발/재건축/기타], type_raw, small_scale).
    소규모주택정비(별법)는 type=기타 + small_scale=True 로 도정법과 분리."""
    r = raw or ""
    if "가로주택" in r:
        return "기타", "가로주택정비", True
    if "소규모재건축" in r:
        return "기타", "소규모재건축", True
    if "소규모재개발" in r:
        return "기타", "소규모재개발", True
    if "자율주택" in r:
        return "기타", "자율주택정비", True
    if "재개발" in r:
        return "재개발", "재개발", False
    if "재건축" in r:
        return "재건축", "재건축", False
    if "주거환경" in r:
        return "기타", "주거환경개선", False
    return "기타", (r.strip() or None), False


def _sigungu_from_loc(loc):
    m = re.search(r"([가-힣]{2,}[구군])", loc or "")
    return m.group(1) if m else None


def _record(name, sigungu, type_raw, area_m2, units_old, units_plan,
            location, executor, status, dates, unparsed, snapshot_only=False):
    typ, traw, small = classify(type_raw)
    rec = {
        "name": (name or "").strip() or None,
        "sigungu": (sigungu or "").strip() or None,
        "type": typ,
        "type_raw": traw,
        "small_scale": small,
        "area_m2": area_m2,
        "units_old": units_old,
        "units_plan": units_plan,
        "location": (location or "").strip() or None,
        "executor": (executor or "").strip() or None,
        "status": (status or "").strip() or None,
        "dates": dates,
    }
    if unparsed:
        rec["unparsed_dates"] = unparsed   # 원문 보존(파싱 불가 단계)
    if snapshot_only:
        rec["snapshot_only"] = True
    return rec


# ─────────────────────────── 대구 ───────────────────────────
def _ensure_xlsx(path: Path) -> Path:
    """OLE(.xls) 이면 soffice 로 .xlsx 변환. 이미 xlsx(PK)면 그대로."""
    head = path.read_bytes()[:8]
    if head[:2] == b"PK":
        return path
    if head[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 = 구식 .xls
        subprocess.run(["soffice", "--headless", "--convert-to", "xlsx",
                        "--outdir", str(path.parent), str(path)],
                       check=True, capture_output=True)
        conv = path.with_suffix(".xlsx")
        if not conv.exists():
            raise RuntimeError(f"soffice 변환 실패: {path}")
        return conv
    raise RuntimeError(f"알 수 없는 파일형식(head={head.hex()}): {path}")


def fetch_daegu(refresh: bool) -> Path:
    """최신 대구 정비사업 추진현황 .xlsx 를 raw 캐시로. 캐시 있으면 재사용."""
    cached = sorted(RAW_DIR.glob("daegu_jeongbi_*.xlsx"),
                    key=lambda p: p.stat().st_mtime, reverse=True)
    if cached and not refresh:
        print(f"  [대구] 캐시 재사용: {cached[0].name}")
        return _ensure_xlsx(cached[0])

    s = requests.Session()
    s.headers.update({"User-Agent": UA})
    r = s.get(DAEGU_LIST, timeout=30)
    r.raise_for_status()
    # '소규모'로 시작하지 않는 최신 '정비사업 추진현황(...)' 게시글(목록은 최신순)
    pairs = re.findall(
        r"fn_icms_navi_common\('view',\s*'(\d+)'\)[^>]*>\s*(정비사업 추진현황[^<]*)</a>",
        r.text)
    if not pairs:
        raise RuntimeError("대구 자료실에서 '정비사업 추진현황' 게시글을 못 찾음")
    nttId, title = pairs[0]
    ym = re.search(r"(\d{4})[.\s]*(\d{1,2})\s*월", title)
    tag = f"{ym.group(1)}-{int(ym.group(2)):02d}" if ym else nttId
    print(f"  [대구] 최신글 nttId={nttId} '{title.strip()}' → {tag}")

    art = s.post("https://www.daegu.go.kr/icms/bbs/selectBoardArticle.do",
                 data={**DAEGU_BBS, "nttId": nttId, "pageIndex": "1"},
                 headers={"Referer": DAEGU_LIST}, timeout=30)
    art.raise_for_status()
    m = re.search(r"fn_egov_downFile\('([^']+)','([^']+)'\)", art.text)
    if not m:
        raise RuntimeError(f"대구 nttId={nttId} 본문에서 첨부(atchFileId) 못 찾음")
    atch, fsn = m.group(1), m.group(2)
    # ⚠ 반드시 /icms/ 접두 컨트롤러(무접두 /cmm/fms 는 한글파일명 인코딩버그로 실패)
    dl = s.get(f"https://www.daegu.go.kr/icms/cmm/fms/FileDown.do"
               f"?atchFileId={atch}&fileSn={fsn}",
               headers={"Referer": "https://www.daegu.go.kr/build/index.do?menu_id=00001338"},
               timeout=60)
    dl.raise_for_status()
    if dl.content[:2] != b"PK" and dl.content[:4] != b"\xd0\xcf\x11\xe0":
        raise RuntimeError(f"대구 다운로드가 스프레드시트가 아님(head={dl.content[:16]!r})")
    out = RAW_DIR / f"daegu_jeongbi_{tag}.xlsx"
    out.write_bytes(dl.content)
    print(f"  [대구] 저장 {out.name} ({len(dl.content):,}B)")
    return _ensure_xlsx(out)


# 대구 9단계 → (년,월,일) 열(1-based). 예비평가·안전진단은 스키마 외(별도 캡처).
_DAEGU_COLS = {
    "예정구역고시": (12, 13, 14), "구역지정": (15, 16, 17), "추진위": (18, 19, 20),
    "조합설립": (27, 28, 29), "사업시행": (30, 31, 32), "관리처분": (33, 34, 35),
    "착공": (36, 37, 38), "일반분양": (39, 40, 41), "준공": (42, 43, 44),
}


def parse_daegu(xlsx: Path) -> list:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    def v(r, c):
        x = ws.cell(r, c).value
        return "" if x is None else str(x).strip()

    out = []
    for r in range(4, ws.max_row + 1):
        gu_no = v(r, 3)                      # 구역번호: '중구-10-11'
        if not re.search(r"-\d", gu_no):     # 소계/공백행 제외
            continue
        dates, unparsed = {}, {}
        for stage, (cy, cm, cd) in _DAEGU_COLS.items():
            iso, raw = parse_ymd(ws.cell(r, cy).value,
                                 ws.cell(r, cm).value, ws.cell(r, cd).value)
            dates[stage] = iso
            if raw:
                unparsed[stage] = raw
        loc = " ".join(p for p in (v(r, 5), v(r, 6)) if p)  # 동 + 지번
        rec = _record(
            name=v(r, 4) or gu_no, sigungu=v(r, 2), type_raw=v(r, 8),
            area_m2=parse_float(v(r, 7)), units_old=None, units_plan=None,
            location=loc, executor=None, status=v(r, 46), dates=dates,
            unparsed=unparsed)
        rec["district_no"] = gu_no
        out.append(rec)
    return out


# ─────────────────────────── 경기 ───────────────────────────
def fetch_gyeonggi(refresh: bool) -> Path:
    out = RAW_DIR / "gyeonggi_sheet.csv"
    if out.exists() and not refresh:
        print(f"  [경기] 캐시 재사용: {out.name}")
        return out
    r = requests.get(GG_CSV, headers={"User-Agent": UA,
                     "Accept-Language": "ko-KR,ko;q=0.9", "Referer": GG_REFERER},
                     timeout=60)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    if "csv" not in ct.lower():
        raise RuntimeError(f"경기 응답이 CSV 아님(content-type={ct})")
    out.write_bytes(r.content)
    print(f"  [경기] 저장 {out.name} ({len(r.content):,}B, {ct})")
    return out


# 경기 CSV 열(0-based)
_GG = dict(sigungu=0, step=1, type=2, name=3, loc=4, area=5,
           units_old=8, units_plan=14, executor=32)
_GG_COLS = {"예정구역고시": 35, "구역지정": 38, "추진위": 40, "조합설립": 43,
            "사업시행": 44, "관리처분": 45, "착공": 46, "일반분양": 47, "준공": 48}


def parse_gyeonggi(csv_path: Path) -> list:
    import csv
    with open(csv_path, encoding="cp949", newline="") as f:
        rows = list(csv.reader(f))
    out = []
    for r in rows[1:]:
        if not any(x.strip() for x in r) or len(r) <= _GG["name"]:
            continue
        if not r[_GG["name"]].strip() and not r[_GG["sigungu"]].strip():
            continue
        dates, unparsed = {}, {}
        for stage, ci in _GG_COLS.items():
            raw = r[ci] if ci < len(r) else ""
            iso, bad = parse_datestr(raw)
            dates[stage] = iso
            if bad:
                unparsed[stage] = bad
        out.append(_record(
            name=r[_GG["name"]], sigungu=r[_GG["sigungu"]], type_raw=r[_GG["type"]],
            area_m2=parse_float(r[_GG["area"]]),
            units_old=parse_int(r[_GG["units_old"]], zero_null=True),
            units_plan=parse_int(r[_GG["units_plan"]], zero_null=True),
            location=r[_GG["loc"]], executor=r[_GG["executor"]],
            status=r[_GG["step"]], dates=dates, unparsed=unparsed))
    return out


# ─────────────────────────── 부산 ───────────────────────────
def fetch_busan(refresh: bool, service_key: str) -> Path:
    out = RAW_DIR / "busan_jeongbi.json"
    if out.exists() and not refresh:
        print(f"  [부산] 캐시 재사용: {out.name}")
        return out
    items, page = [], 1
    total = None
    while True:
        r = requests.get(BUSAN_API, params={
            "serviceKey": service_key, "pageNo": page,
            "numOfRows": 1000, "resultType": "json"},
            headers={"User-Agent": UA}, timeout=60)
        r.raise_for_status()
        try:
            j = r.json()
        except ValueError:
            raise RuntimeError(f"부산 API 비-JSON 응답(HTTP {r.status_code}): {r.text[:200]}")
        hdr = j["response"]["header"]
        if hdr.get("resultCode") not in ("00", "0"):
            raise RuntimeError(f"부산 API 오류: {hdr}")
        body = j["response"]["body"]
        total = int(body.get("totalCount", 0))
        batch = body.get("items", {}).get("item", []) or []
        if isinstance(batch, dict):
            batch = [batch]
        items.extend(batch)
        if len(items) >= total or not batch:
            break
        page += 1
    out.write_text(json.dumps({"totalCount": total, "items": items},
                              ensure_ascii=False), encoding="utf-8")
    print(f"  [부산] 저장 {out.name} ({len(items)}/{total}건)")
    return out


def parse_busan(json_path: Path) -> list:
    d = json.loads(json_path.read_text(encoding="utf-8"))
    items = d["items"] if "items" in d else d["response"]["body"]["items"]["item"]
    null_dates = {s: None for s in SCHEMA}
    out = []
    for i in items:
        name = i.get("areaName")
        loc = i.get("location")
        out.append(_record(
            name=name, sigungu=_sigungu_from_loc(loc), type_raw=name,
            area_m2=parse_float(i.get("areaUnit")), units_old=None,
            units_plan=parse_int(i.get("generationJoo"), zero_null=True),
            location=loc, executor=i.get("businessEntities"),
            status=i.get("step"), dates=dict(null_dates), unparsed=None,
            snapshot_only=True))
    return out


# ─────────────────────────── 검증 ───────────────────────────
def _median(xs):
    xs = sorted(xs)
    n = len(xs)
    if not n:
        return None
    return xs[n // 2] if n % 2 else (xs[n // 2 - 1] + xs[n // 2]) / 2


def validate(regions: dict):
    print("\n" + "=" * 68)
    print("검증 출력")
    print("=" * 68)
    for reg, recs in regions.items():
        n = len(recs)
        small = sum(1 for r in recs if r["small_scale"])
        snap = any(r.get("snapshot_only") for r in recs)
        print(f"\n[{reg}] 구역 {n}건 (도정법 {n - small} · 소규모 {small})"
              + ("  ※ snapshot_only(이력일자 없음)" if snap else ""))
        # 단계별 보유율
        print("  단계일자 보유율:")
        for s in SCHEMA:
            c = sum(1 for r in recs if r["dates"].get(s))
            print(f"    {s:<8} {c:>4}/{n}  {100*c/n:5.1f}%")
        # 소요기간 새니티: 조합설립 → 사업시행 (중위 연수 1~10 기대)
        gaps = []
        for r in recs:
            a, b = r["dates"].get("조합설립"), r["dates"].get("사업시행")
            if a and b:
                da = _dt.date.fromisoformat(a)
                db = _dt.date.fromisoformat(b)
                gaps.append((db - da).days / 365.25)
        med = _median([g for g in gaps if g >= 0])
        if med is not None:
            ok = "OK" if 1 <= med <= 10 else "⚠ 범위벗어남"
            print(f"  소요기간 조합설립→사업시행: n={len(gaps)} 중위 {med:.2f}년  [{ok}]")
        else:
            print(f"  소요기간 조합설립→사업시행: 표본 없음(snapshot 또는 미도달)")


# ─────────────────────────── main ───────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--refresh", action="store_true", help="원본 재다운로드(캐시 무시)")
    args = ap.parse_args()

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    cfg = json.loads(CONFIG.read_text())
    service_key = cfg["service_key"]

    print("정비사업 Phase 1 수집 시작")
    daegu_x = fetch_daegu(args.refresh)
    gg_c = fetch_gyeonggi(args.refresh)
    busan_j = fetch_busan(args.refresh, service_key)

    regions = {
        "대구": parse_daegu(daegu_x),
        "경기": parse_gyeonggi(gg_c),
        "부산": parse_busan(busan_j),
    }

    payload = {
        "schema": SCHEMA,
        "regions": regions,
        "meta": {
            "generated": _dt.datetime.now().strftime("%Y-%m-%d"),
            "counts": {k: len(v) for k, v in regions.items()},
            "sources": {
                "대구": "daegu.go.kr/build 자료실(BBS_00153) — 정비사업 추진현황 .xlsx (원 출처, 도정법 9단계 표준)",
                "경기": "data.gg.go.kr sheet downloadSheetData.do infId=S62GFEEN7JMLMA0PH6CF19108891 (CP949 CSV)",
                "부산": "apis.data.go.kr/6260000/MaintenanceBusinessStatus1/getMaintenanceBusiness1 (스냅샷, snapshot_only)",
            },
            "notes": [
                "type=재개발/재건축/기타. 주거환경(개선)·소규모정비는 기타. small_scale=True 는 별법(빈집·소규모주택정비특례법).",
                "status=스냅샷 추진단계(대구 사업추진단계 / 경기 사업단계 / 부산 step). dates=이력 인가일자.",
                "부산은 이력 일자 미제공 → dates 전부 null, snapshot_only=true.",
                "unparsed_dates: 원문이 있으나 완전 파싱 불가한 단계의 원문 보존(예: 연도만 있는 대구 구역지정).",
            ],
        },
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\n산출: {OUT.relative_to(ROOT)} "
          f"({sum(len(v) for v in regions.values())} 구역)")

    validate(regions)


if __name__ == "__main__":
    main()
